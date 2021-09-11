import openpyxl

inv_file = openpyxl.load_workbook('excel/practice_excel.xlsx')
working_sheet = inv_file['Sheet1']

products_per_supply = {}
total_value_per_supplier = {}
product_under_10_inv = {}

for prd_row in range(2, working_sheet.max_row + 1):
    supplier_name = working_sheet.cell(prd_row, 4).value
    price = working_sheet.cell(prd_row, 3).value
    inventory = working_sheet.cell(prd_row, 2).value
    product_num = working_sheet.cell(prd_row, 1).value
    inv_price = working_sheet.cell(prd_row, 5)

    # calculate number of products per supplier
    if supplier_name in products_per_supply:
        products_per_supply[supplier_name] = products_per_supply[supplier_name] + 1
    else:
        print('adding new products')
        products_per_supply[supplier_name] = 1

    # calculate number of inventory per supplier
    if supplier_name in total_value_per_supplier:
        current_total_value = total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[supplier_name] = current_total_value + inventory * price

    else:
        total_value_per_supplier[supplier_name] = inventory * price

    # calculate number of products with inventory less than 10
    if inventory < 10:
        product_under_10_inv[int(product_num)] = int(inventory)

    # add value for total inventory price
    inv_price.value = inventory * price

inv_file.save("inventory_with_total_value.xlsx")
