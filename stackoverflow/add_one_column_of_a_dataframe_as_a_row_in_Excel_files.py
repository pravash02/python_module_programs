import pandas as pd
from openpyxl import load_workbook
from pandas import ExcelWriter

old_df = pd.read_excel('test.xlsx', engine='openpyxl')

new_df = pd.DataFrame({
    'A': [1, 0, 3, 0, 5],
    'Company': [0, 2, 0, 4, 0],
    'C': [6, 7, 0, 9, 10]
})

# res = old_df.append(new_df, ignore_index=True)
updated_df = pd.concat([old_df, new_df], ignore_index=True)
writer = ExcelWriter('test.xlsx')
updated_df['Company'].to_excel(writer, 'Sheet1', index=False)
writer.save()


# wb = load_workbook('file_test.xlsx')
# ws = wb.active
# ws.insert_cols(1)
# ws['A1'] = 'Company'
# for j, row in df.iterrows():
#     ws.cell(row=j+2, column=1).value = row['Company']
# wb.save('file_test.xlsx')


df = pd.read_excel('File1.xlsx', engine='openpyxl')

new_row = pd.Series(['Company XYZ', 1000], index=df.columns)

df = df.append(new_row, ignore_index=True)

df.to_excel('File1.xlsx', 'Sheet1', index=False)